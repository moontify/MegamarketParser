from aiogram import Bot, Dispatcher
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
import asyncio
import aiohttp
from openpyxl import load_workbook, Workbook
from datetime import datetime
from pydrive.auth import GoogleAuth
from pydrive.drive import GoogleDrive
import os
import json
import ast
import configparser
import re
word = input()
TELEGRAM_TOKEN = "Ваш токен бота"
bot = Bot(token=TELEGRAM_TOKEN)
dp = Dispatcher(bot=bot)
CHAT_ID = "Ваш чат ид"
processed_products_file = 'processed_products.txt'

gauth = GoogleAuth()
gauth.LocalWebserverAuth()
drive = GoogleDrive(gauth)

def is_product_processed(product_id):
    with open(processed_products_file, 'r') as file:
        processed_ids = file.read().splitlines()
    return product_id in processed_ids

def mark_product_as_processed(product_id):
    with open(processed_products_file, 'a') as file:
        file.write(f'{product_id}\n')

async def send_message(chat_id, text, parse_mode='HTML'):
    await bot.send_message(chat_id, text, parse_mode=parse_mode)


def append_to_xlsx(filename, data, sheet_name='Sheet1'):
    if not data:
        return

    if os.path.exists(filename):
        workbook = load_workbook(filename)
        sheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.create_sheet(sheet_name)
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = sheet_name
        headers = ['Дата и время', 'Название', 'Цена', 'Бонусы %', 'Бонусы ₽', 'Цена с вычетом бонусов', 'Продавец', 'Ссылка']
        sheet.append(headers)


    last_row = sheet.max_row
    if last_row > 1:
        last_row += 3

    for item in data:
        current_datetime = datetime.now()
        row = [current_datetime, item['name'], item['price'], item['bonus_percent'], item['bonus_value'], item['lastcost'],
               item['store_name'], item['link']]
        sheet.append(row)

    workbook.save(filename)

async def main():
    path_to_chromedriver = 'chromedriver-win64/chromedriver.exe'
    service = Service(executable_path=path_to_chromedriver)
    #user_agent = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36'
    options = webdriver.ChromeOptions()
    options.add_argument('--start-maximized')
    options.add_argument('--blink-settings=imagesEnabled=true')
    #options.add_argument(f'user-agent={user_agent}')
    #options.add_argument("--incognito")
    options.add_experimental_option("prefs", {"profile.managed_default_content_settings.images": 2})
    driver = webdriver.Chrome(service=service, options=options)
    index = 1

    # Вынужденная прогрузка из за капчи(
    driver.get('https://megamarket.ru/')
    def config_read(index):
        config = configparser.ConfigParser()
        config.read('settings.cfg')
        settings_str = config['Settings' + '_' + word]['settings']
        settings_list = ast.literal_eval(f'[{settings_str}]')
        if index <= len(settings_list):
            settings = settings_list[index - 1]
            return {
                'link': str(settings['link' + str(index) + word]),
                'min_bonus_percent': float(settings['bonus' + str(index) + word]),
                'min_price': float(settings['minPrice' + str(index) + word]),
                'max_price': float(settings['maxPrice' + str(index) + word]),
                'workafrika': bool(settings['work247' + str(index) + word]),
                'addToTable': bool(settings['addToTable' + str(index) + word]),
                'addTogDrive': bool(settings['addTogDrive' + str(index) + word])
            }
        else:
            return None

    products = []
    page_number = 1

    while True:
        page_settings = config_read(index)
        if page_settings is None:
            index = 1
            page_number = 1
            print("ок")
            continue
        page_settings = config_read(index)
        workafrika = page_settings['workafrika']
        try:
            if workafrika == True:
                while True:
                        page_settings = config_read(index)
                        if page_settings is None:
                            index = 1
                            page_number = 1
                            continue
                        page_settings = config_read(index)
                        links = page_settings['link']
                        min_bonus_percent = page_settings['min_bonus_percent']
                        min_price = page_settings['min_price']
                        max_price = page_settings['max_price']
                        addToTable = page_settings['addToTable']
                        addTogDrive = page_settings['addTogDrive']

                        out_stock = await run_selenium(driver, links, CHAT_ID, min_bonus_percent, min_price, max_price,page_number, products, index, config_read,addToTable, addTogDrive)
                        if addTogDrive == True:
                            def create_and_upload_file(file_path='products.xlsx'):
                                try:
                                    file_title = os.path.basename(file_path)
                                    my_file = drive.CreateFile({'title': file_title})

                                    my_file.SetContentFile(file_path)
                                    my_file.Upload()

                                    return f'Файл {file_title} успешно загружен!'
                                except Exception as _ex:
                                    return f'Ошибка при загрузке {_ex}'

                            file_path = 'products.xlsx'
                            print(create_and_upload_file(file_path=file_path))
                        if out_stock:
                            await send_message(CHAT_ID,'Страницы кончились, переключаюсь на следующую ссылку')
                            index += 1
                            page_number = 1
                        else:
                            page_number += 1






        finally:
            driver.quit()





sent_messages_ids = []
async def run_selenium(driver, links, chat_id, min_bonus_percent, min_price, max_price, page_number, products, index, config_read, addToTable, addTogDrive):
    await asyncio.sleep(5)
    index = links.find('#?filters')
    out = False
    if index != -1:
        page_url = links[:index] + f'/page-{page_number}/' + links[index:]
        print(page_url)
    else:
        print("Не найдено '#?filters' в URL.")
        print(links + f'/page-{page_number}/')
        page_url = links[:index] + f'/page-{page_number}/'
    driver.get(page_url)
    cookies_file_path = 'package.json'



    #На линус при добавлении куки появляется капча
    #Они защиту поменяли :)))


    with open(cookies_file_path, 'r') as cookies_file:
        cookies = json.load(cookies_file)
        with open(cookies_file_path, 'r') as cookies_file:
            cookies = json.load(cookies_file)

            for cookie in cookies:
                if 'expiry' in cookie:
                    del cookie['expiry']
                if 'sameSite' not in cookie or cookie['sameSite'] not in ["Strict", "Lax", "None"]:
                    cookie['sameSite'] = "Lax"
                driver.add_cookie(cookie)
            driver.refresh()
    await asyncio.sleep(10)
    try:
        out_of_stock = driver.find_element(By.CSS_SELECTOR, ".catalog-items-list__out-of-stock-heading")
        if out_of_stock is not None:
            out = True
            return out
    except NoSuchElementException:
        print(page_number)
        out = False
    product_elements = driver.find_elements(By.CSS_SELECTOR, 'div[data-product-id]')
    for product in product_elements:
        product_info = {}
        product_name_element = product.find_element(By.CSS_SELECTOR, 'a.catalog-item-regular-desktop__title-link')
        product_info['name'] = product_name_element.get_attribute('title')
        print(product_info['name'])
        store_name_elements = product.find_elements(By.CSS_SELECTOR, 'span.merchant-info__name')
        product_info['store_name'] = store_name_elements[0].text if store_name_elements else 'Store name not available'
        product_count = driver.find_element(By.CSS_SELECTOR, '[itemprop="offerCount"]').get_attribute('content')
        price_element = product.find_element(By.CSS_SELECTOR, 'div.catalog-item-regular-desktop__price')
        price_text = price_element.text.replace(' ', '').replace('₽', '').replace(',', '.')
        product_info['price'] = float(re.sub(r'\D', '', price_text))
        bonus_percent_elements = product.find_elements(By.CSS_SELECTOR, 'span.bonus-percent')
        product_info['bonus_percent'] = int(
            bonus_percent_elements[0].text.replace('%', '')) if bonus_percent_elements else 0
        bonus_elements = product.find_elements(By.CSS_SELECTOR, 'span.bonus-amount')
        bonus_text = bonus_elements[0].text.replace(' ', '').replace('₽', '').replace(',',
                                                                                              '.') if bonus_elements else '0'
        product_info['bonus_value'] = float(re.sub(r'\D', '', bonus_text))
        product_info['lastcost'] = product_info['price'] - product_info['bonus_value']
        link_element = product.find_element(By.CSS_SELECTOR, "a.catalog-item-regular-desktop__title-link")
        product_info['link'] = link_element.get_attribute('href')
        message_id = product_info['link']
        if message_id not in sent_messages_ids:
            if product_info['bonus_percent'] >= min_bonus_percent and min_price <= product_info['price'] <= max_price:
                products.append(product_info)

                message_text = (
                    f"<b>{product_info['name']}</b>\n"
                    f"<b>Цена:</b> {product_info['price']}₽\n"
                    f"<b>Бонусы:</b> {product_info['bonus_percent']}% ({product_info['bonus_value']}₽)\n"
                    f"<b>Цена с вычетом бонусов:</b> {product_info['lastcost']}₽\n"
                    f"<b>Продавец:</b> {product_info['store_name']}\n"
                    f"<a href='{product_info['link']}'>Ссылка на товар</a>"
                    )
                sent_messages_ids.append(message_id)
                await send_message(chat_id, message_text)
        else:
            print("Леее такое уже было")
            continue

        if addToTable == True and message_id not in sent_messages_ids:
            if products:
                    append_to_xlsx("products.xlsx", products)


    return out





if __name__ == '__main__':
    asyncio.run(main())